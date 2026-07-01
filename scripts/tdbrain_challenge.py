"""TDBRAIN-Challenge submission driver (train on Discovery -> predict on blinded Replication).

Brainclinics' blinded out-of-sample challenge: predict a phenotype on the held-out
Replication set and email the filled template to martijn@brainclinics.com. Targets:
  diagnostic  -- depression (MDD vs non-MDD, Balanced Accuracy >60%), age, sex
  prognostic  -- treatment response/remission (nPPV >110%)

Self-contained: resting BDF/BDF+ (V3.1) or BrainVision (legacy) -> 26 scalp channels ->
per-channel log band-power + aperiodic (1/f) slope; RidgeCV(+bias) for age, balanced
LogisticRegression for MDD/sex/treatment; optional LEACE deconfounding of site/age (our
"identity-free" differentiator). Fills only the prediction column of the provided template.

V3.1 paths + datasheet columns + template format are CLI-parameterized (finalize when the
password-unzipped V3.1 lands). ``--self-cv`` smoke-tests the ML core on a labeled cohort.
Usage:
  # smoke test on the 22-subj sample (LOO age MAE + sex BA):
  tdbrain_challenge.py --self-cv --bids /mnt/t9/tdbrain/bids \
      --participants /mnt/t9/tdbrain/.../participants.tsv --target age
  # real submission (once V3.1 staged):
  tdbrain_challenge.py --target mdd --features bandpower --deconfound site \
      --discovery-bids ... --discovery-labels datasheet.xlsx \
      --replication-bids ... --template diagnostic_template.xlsx --out submission.xlsx
"""
from __future__ import annotations

import argparse
import glob
import os
import sys

import numpy as np

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from emeg_fm.tdbrain_cohort import (  # noqa: E402
    TDBRAIN_SCALP_26, _find_raw, _read_raw, _sid_to_int,
)

BANDS = {"delta": (1, 4), "theta": (4, 8), "alpha": (8, 13), "beta": (13, 30), "gamma": (30, 45)}


# ---------- features ----------
def recording_features(path, channels=TDBRAIN_SCALP_26, sfreq_target=200.0):
    """Per-channel log band-power (5 bands) + aperiodic 1/f slope for one resting file.

    Returns a (26*6,) feature vector, or None if the montage is incomplete / too short.
    """
    from scipy.signal import welch

    raw = _read_raw(path)
    present = [c for c in channels if c in raw.ch_names]
    if len(present) != len(channels):
        return None
    raw.pick(present)
    if raw.info["sfreq"] > sfreq_target:
        raw.resample(sfreq_target, verbose="error")
    x = raw.get_data()  # (26, T), Volts
    sf = raw.info["sfreq"]
    if x.shape[1] < int(4 * sf):
        return None
    f, pxx = welch(x, fs=sf, nperseg=int(2 * sf), axis=1)  # (26, F)
    feats = []
    for lo, hi in BANDS.values():
        m = (f >= lo) & (f < hi)
        feats.append(np.log(pxx[:, m].mean(axis=1) + 1e-30))  # (26,)
    # aperiodic slope: loglog fit of PSD over 2-40 Hz per channel
    band = (f >= 2) & (f <= 40)
    lf, lp = np.log(f[band] + 1e-12), np.log(pxx[:, band] + 1e-30)
    slope = np.polyfit(lf, lp.T, 1)[0]  # (26,)
    feats.append(slope)
    return np.concatenate(feats).astype(np.float64)  # (26*6,)


def build_matrix(bids_root, subject_ids, task="restEC"):
    """(X (n, d), kept_ids) for the given subjects; drops unloadable recordings."""
    X, kept = [], []
    for sid in subject_ids:
        p = _find_raw(bids_root, sid, task)
        if p is None:
            continue
        v = recording_features(p)
        if v is None:
            continue
        X.append(v)
        kept.append(sid)
    return (np.vstack(X) if X else np.empty((0, 0))), kept


# ---------- labels ----------
def read_labels(path, id_col="participant_id"):
    """Read participants.tsv / datasheet (.tsv/.csv/.xlsx) -> {int sid: {col: val}}."""
    rows = _read_table(path)
    out = {}
    for r in rows:
        sid = _sid_to_int(str(r.get(id_col, "")))
        if sid is not None and sid not in out:
            out[sid] = r
    return out


def _read_table(path):
    if path.lower().endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(it)]
        return [dict(zip(header, row)) for row in it]
    import csv
    delim = "\t" if path.lower().endswith(".tsv") else ","
    with open(path, newline="") as f:
        return list(csv.DictReader(f, delimiter=delim))


# ---------- model ----------
def _target_vector(labels, sids, target, spec):
    """(y array aligned to sids, is_regression) for the requested target."""
    col = spec[target]["col"]
    if target == "age":
        y = np.array([float(labels[s][col]) for s in sids])
        return y, True
    pos = spec[target]["positive"]
    y = np.array([1 if str(labels[s].get(col, "")).strip().upper() in pos else 0 for s in sids])
    return y, False


def fit_predict(Xtr, ytr, Xte, is_reg, deconfound=None, conf_tr=None, conf_te=None):
    from sklearn.linear_model import LogisticRegression, RidgeCV
    from sklearn.preprocessing import StandardScaler

    sc = StandardScaler().fit(Xtr)
    Xtr, Xte = sc.transform(Xtr), sc.transform(Xte)
    if deconfound is not None and conf_tr is not None:
        sys.path.insert(0, os.path.expanduser("~/dev/meeg-brain-age-benchmark-paper"))
        from leace import LeaceEraser  # closed-form LEACE: fit(X, site) + transform(X)
        er = LeaceEraser().fit(Xtr, conf_tr)
        Xtr, Xte = er.transform(Xtr), er.transform(Xte)
    if is_reg:
        m = RidgeCV(alphas=np.logspace(-3, 5, 25)).fit(Xtr, ytr)
        return m.predict(Xte)
    m = LogisticRegression(max_iter=2000, class_weight="balanced").fit(Xtr, ytr)
    return m.predict_proba(Xte)[:, 1]


def self_cv(X, y, is_reg):
    """Leave-one-out internal metric on the labeled cohort (MAE or Balanced Accuracy)."""
    from sklearn.metrics import balanced_accuracy_score, mean_absolute_error
    from sklearn.model_selection import LeaveOneOut
    pred = np.zeros(len(y), float)
    for tr, te in LeaveOneOut().split(X):
        pred[te] = fit_predict(X[tr], y[tr], X[te], is_reg)
    if is_reg:
        return "MAE", mean_absolute_error(y, pred)
    return "BalancedAcc", balanced_accuracy_score(y, (pred >= 0.5).astype(int))


TARGET_SPEC = {
    "age": {"col": "age"},
    "sex": {"col": "gender", "positive": {"1", "M", "MALE"}},
    "mdd": {"col": "indication", "positive": {"MDD", "DEPRESSION"}},
    "treatment": {"col": "Responder", "positive": {"1", "RESPONDER", "REMITTER", "R"}},
}


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--target", choices=list(TARGET_SPEC), default="age")
    ap.add_argument("--task", default="restEC")
    ap.add_argument("--self-cv", action="store_true", help="internal LOO metric on one labeled cohort")
    ap.add_argument("--bids"); ap.add_argument("--participants")
    ap.add_argument("--discovery-bids"); ap.add_argument("--discovery-labels")
    ap.add_argument("--replication-bids"); ap.add_argument("--template"); ap.add_argument("--out")
    ap.add_argument("--deconfound", default=None)
    a = ap.parse_args()

    if a.self_cv:
        labels = read_labels(a.participants)
        col = TARGET_SPEC[a.target]["col"]
        sids = [s for s in labels if str(labels[s].get(col, "")).strip() not in ("", "n/a", "None")]
        X, kept = build_matrix(a.bids, sids, a.task)
        if len(kept) < 5:
            sys.exit(f"only {len(kept)} loadable labeled recordings — need >=5")
        labels_k = {s: labels[s] for s in kept}
        y, is_reg = _target_vector(labels_k, kept, a.target, TARGET_SPEC)
        name, val = self_cv(X, y, is_reg)
        print(f"[self-cv] target={a.target} n={len(kept)} feat={X.shape[1]} | {name}={val:.3f}")
        if not is_reg:
            print(f"          class balance: {int(y.sum())}/{len(y)} positive")
        return

    # real submission: Discovery train -> Replication predict -> fill template
    disc_labels = read_labels(a.discovery_labels)
    col = TARGET_SPEC[a.target]["col"]
    disc_sids = [s for s in disc_labels if str(disc_labels[s].get(col, "")).strip() not in ("", "n/a")]
    Xtr, disc_kept = build_matrix(a.discovery_bids, disc_sids, a.task)
    ytr, is_reg = _target_vector({s: disc_labels[s] for s in disc_kept}, disc_kept, a.target, TARGET_SPEC)
    repl_sids = _template_subject_ids(a.template)
    Xte, repl_kept = build_matrix(a.replication_bids, repl_sids, a.task)
    preds = fit_predict(Xtr, ytr, Xte, is_reg)
    fill_template(a.template, dict(zip(repl_kept, preds)), a.out)
    name, val = self_cv(Xtr, ytr, is_reg)
    print(f"[submission] target={a.target} discovery_n={len(disc_kept)} replication_n={len(repl_kept)}")
    print(f"             internal {name}={val:.3f} | wrote {a.out}")


def _template_subject_ids(template):
    rows = _read_table(template)
    ids = []
    for r in rows:
        for k in ("participant_id", "subject", "ID", "id"):
            if k in r and r[k]:
                sid = _sid_to_int(str(r[k]))
                if sid is not None:
                    ids.append(sid)
                break
    return ids


def fill_template(template, sid_to_pred, out):
    """Fill ONLY the prediction column of the provided template (keep everything else)."""
    import shutil
    shutil.copy(template, out)
    if out.lower().endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        header = [c.value for c in ws[1]]
        idc = next((i for i, h in enumerate(header) if str(h).lower() in ("participant_id", "subject", "id")), 0)
        pc = next((i for i, h in enumerate(header) if "predict" in str(h).lower()), len(header) - 1)
        for row in ws.iter_rows(min_row=2):
            sid = _sid_to_int(str(row[idc].value))
            if sid in sid_to_pred:
                row[pc].value = float(sid_to_pred[sid])
        wb.save(out)


if __name__ == "__main__":
    main()
